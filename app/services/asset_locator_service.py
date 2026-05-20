"""Asset locator: pick assets by model token(s) and group by network.

Used by the "where are my ThinkBooks?" lookup panel. Match is substring,
case-insensitive, against the effective model fields (override_model first,
then series, then raw model) — so picking "ThinkBook" catches all variants
like "ThinkBook 14 G6 ABP" without the user having to know SKU strings."""

from __future__ import annotations

import csv
import io
import ipaddress
from dataclasses import dataclass, field
from datetime import datetime
from typing import Iterable

from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.models.inventory import Asset, Network


# Asset types where `network_id` is the authoritative answer (Meraki gear
# linked by serial → networkId). For everything else we re-evaluate live
# from defender_last_ip, since the FK is just a cache that goes stale
# whenever Defender IPs change between Meraki syncs.
_GEAR_ASSET_TYPES = {"gateway", "switch", "ap"}


@dataclass
class SeenNetworkRef:
    """One 'this MAC was seen on network X' entry — drives the locator's
    'Networks seen' column and the seen-networks CSV column."""
    network_id: int
    network_name: str
    last_seen_at: datetime | None
    ip: str | None
    vlan: int | None


@dataclass
class LocatorDeviceRow:
    asset_id: int
    serial_number: str
    asset_type: str
    manufacturer: str | None
    model: str | None
    override_model: str | None
    series: str | None
    generation: str | None
    effective_model: str
    os: str | None
    os_version: str | None
    status_code: str
    assigned_upn: str | None
    location_name: str | None
    defender_last_ip: str | None
    mac_address: str | None
    intune_device_name: str | None
    network_id: int | None
    network_name: str | None
    network_subnet: str | None
    matched_vlan_id: int | None
    matched_vlan_name: str | None
    matched_token: str
    seen_networks: list[SeenNetworkRef] = field(default_factory=list)

    @property
    def seen_networks_str(self) -> str:
        """Comma-joined network names for CSV/XLSX export. Empty when no
        cached appearances exist."""
        return ", ".join(s.network_name for s in self.seen_networks)


@dataclass
class LocatorGroup:
    network_id: int | None
    network_name: str | None
    network_subnet: str | None
    devices: list[LocatorDeviceRow] = field(default_factory=list)


@dataclass
class LocatorResult:
    tokens: list[str]
    matched: int
    groups: list[LocatorGroup]


# Frontend-shared columns for export — kept here so JSON/CSV/XLSX stay in
# sync. Pairs are (header_label, attribute_name).
EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("Asset ID", "asset_id"),
    ("Serial number", "serial_number"),
    ("Asset type", "asset_type"),
    ("Manufacturer", "manufacturer"),
    ("Model", "effective_model"),
    ("Raw model", "model"),
    ("Series", "series"),
    ("Generation", "generation"),
    ("Override model", "override_model"),
    ("OS", "os"),
    ("OS version", "os_version"),
    ("Status", "status_code"),
    ("Assigned UPN", "assigned_upn"),
    ("Location", "location_name"),
    ("Intune device", "intune_device_name"),
    ("Defender last IP", "defender_last_ip"),
    ("Network", "network_name"),
    ("Subnet", "network_subnet"),
    ("MAC", "mac_address"),
    ("Matched VLAN ID", "matched_vlan_id"),
    ("Matched VLAN name", "matched_vlan_name"),
    ("Matched token", "matched_token"),
    ("Networks seen on (Meraki)", "seen_networks_str"),
]


def _effective_model(a: Asset) -> str:
    if a.override_model and a.override_model.strip():
        return a.override_model.strip()
    parts = [a.series, a.generation]
    joined = " ".join(p for p in parts if p)
    if joined and joined.strip() and joined.strip() != (a.model or "").strip():
        return joined.strip()
    return (a.model or "").strip() or "—"


def _match_token(a: Asset, token_lc: str) -> bool:
    for field_val in (a.override_model, a.series, a.model):
        if field_val and token_lc in field_val.lower():
            return True
    return False


def locate(
    db: Session,
    tokens: Iterable[str],
    *,
    include_archived: bool = False,
) -> LocatorResult:
    """Run the lookup. Returns flat rows grouped into network buckets,
    with unassigned/no-network rows in a trailing `network_id=None` bucket."""

    clean_tokens = [t.strip() for t in tokens if t and t.strip()]
    if not clean_tokens:
        return LocatorResult(tokens=[], matched=0, groups=[])
    lc_tokens = [t.lower() for t in clean_tokens]

    # Pull all candidate assets in one query. Filtering happens in Python
    # because matching against three fields with case-insensitive contains
    # is awkward in SQLite's collation; the row count is small (~thousands
    # max for an org) so the round-trip cost is negligible.
    stmt = select(Asset)
    if not include_archived:
        stmt = stmt.where(Asset.archived_at.is_(None))
    # Cheap pre-filter — at least one of the candidate columns must be
    # non-null. Real matching still happens in Python.
    stmt = stmt.where(
        or_(
            Asset.override_model.is_not(None),
            Asset.series.is_not(None),
            Asset.model.is_not(None),
        )
    )
    assets = list(db.execute(stmt).scalars())

    # Build shared VLAN index — same one the asset linker uses. Cheap
    # SQL+parse pass, run once per request.
    from app.services import meraki_client_service, network_service

    vlan_index = network_service.build_vlan_index(db)
    nets_by_id: dict[int, Network] = {}
    if vlan_index:
        # Pre-fetch Network rows referenced by the index so we can show
        # display_name + cached corp/default cidr fields without N+1 loads.
        ids = {s.network_id for s in vlan_index}
        nets_by_id = {
            n.id: n
            for n in db.execute(
                select(Network).where(Network.id.in_(ids))
            ).scalars()
        }
    # Also load networks that gear is FK'd to, so gear (no VLAN match) can
    # still surface a network name.
    gear_net_ids = {
        a.network_id for a in assets if a.network_id and a.network_id not in nets_by_id
    }
    if gear_net_ids:
        for n in db.execute(
            select(Network).where(Network.id.in_(gear_net_ids))
        ).scalars():
            nets_by_id[n.id] = n

    def _resolve(a: Asset) -> tuple[int | None, network_service.VlanSubnet | None]:
        """Returns (network_id, matched_vlan). Gear keeps its cached FK
        (Meraki serial is authoritative, no matching needed). Clients
        resolve live from defender_last_ip → most-specific VLAN match."""
        if a.asset_type in _GEAR_ASSET_TYPES:
            return (a.network_id, None)
        hit = network_service.match_ip(a.defender_last_ip or "", vlan_index)
        if hit is None:
            return (None, None)
        return (hit.network_id, hit)

    # First pass: figure out which assets match a token. We need the
    # full asset list to do a bulk MAC → appearances fetch before building
    # rows (cheaper than N+1 lookups per row).
    matched_assets_with_token: list[tuple[Asset, str]] = []
    for a in assets:
        hit_token: str | None = None
        for tok_lc, tok_raw in zip(lc_tokens, clean_tokens):
            if _match_token(a, tok_lc):
                hit_token = tok_raw
                break
        if hit_token is not None:
            matched_assets_with_token.append((a, hit_token))

    macs = [a.mac_address for a, _ in matched_assets_with_token if a.mac_address]
    appearances_by_mac = (
        meraki_client_service.appearances_for_macs(db, macs) if macs else {}
    )

    matched_rows: list[LocatorDeviceRow] = []
    for a, hit_token in matched_assets_with_token:
        nid, vlan_hit = _resolve(a)
        n = nets_by_id.get(nid) if nid else None
        if vlan_hit:
            subnet_for_display = vlan_hit.cidr
        elif n:
            subnet_for_display = n.corp_vlan_subnet or n.subnet_cidr
        else:
            subnet_for_display = None
        seen = appearances_by_mac.get(a.mac_address, []) if a.mac_address else []
        seen_refs = [
            SeenNetworkRef(
                network_id=s.network_id,
                network_name=s.network_name,
                last_seen_at=s.last_seen_at,
                ip=s.ip,
                vlan=s.vlan,
            )
            for s in seen
        ]
        matched_rows.append(
            LocatorDeviceRow(
                asset_id=a.id,
                serial_number=a.serial_number,
                asset_type=a.asset_type,
                manufacturer=a.manufacturer,
                model=a.model,
                override_model=a.override_model,
                series=a.series,
                generation=a.generation,
                effective_model=_effective_model(a),
                os=a.os,
                os_version=a.os_version,
                status_code=a.status_code,
                assigned_upn=a.assigned_upn,
                location_name=a.location.name if a.location else None,
                defender_last_ip=a.defender_last_ip,
                mac_address=a.mac_address,
                intune_device_name=a.intune_device_name,
                network_id=nid,
                network_name=(n.display_name if n else None),
                network_subnet=subnet_for_display,
                matched_vlan_id=vlan_hit.vlan_id if vlan_hit else None,
                matched_vlan_name=vlan_hit.vlan_name if vlan_hit else None,
                matched_token=hit_token,
                seen_networks=seen_refs,
            )
        )

    # Group by network. None bucket last.
    by_net: dict[int | None, LocatorGroup] = {}
    for row in matched_rows:
        key = row.network_id
        grp = by_net.get(key)
        if grp is None:
            grp = LocatorGroup(
                network_id=row.network_id,
                network_name=row.network_name,
                network_subnet=row.network_subnet,
            )
            by_net[key] = grp
        grp.devices.append(row)

    # Sort: networks alphabetically by name, no-network bucket last; devices
    # inside each group by effective model + serial.
    groups = sorted(
        by_net.values(),
        key=lambda g: (g.network_id is None, (g.network_name or "").lower()),
    )
    for g in groups:
        g.devices.sort(
            key=lambda d: (d.effective_model.lower(), d.serial_number)
        )

    return LocatorResult(
        tokens=clean_tokens,
        matched=len(matched_rows),
        groups=groups,
    )


# ──────────────────────────── Export helpers ────────────────────────────


def _filename_stub(tokens: list[str]) -> str:
    if not tokens:
        return "asset-locator"
    stamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    slug = "-".join(t.lower().replace(" ", "_") for t in tokens[:3])
    return f"asset-locator_{slug}_{stamp}"


def to_csv_bytes(result: LocatorResult) -> tuple[bytes, str]:
    """Serialise the result as CSV. Returns (body, filename)."""
    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    writer.writerow([h for h, _ in EXPORT_COLUMNS])
    for g in result.groups:
        for d in g.devices:
            writer.writerow(
                [_render_cell(getattr(d, attr)) for _, attr in EXPORT_COLUMNS]
            )
    return buf.getvalue().encode("utf-8"), f"{_filename_stub(result.tokens)}.csv"


def to_xlsx_bytes(result: LocatorResult) -> tuple[bytes, str]:
    """Serialise the result as XLSX. Returns (body, filename)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Locator"

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF1F2937")
    for i, (h, _) in enumerate(EXPORT_COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")

    row_idx = 2
    for g in result.groups:
        for d in g.devices:
            for col_idx, (_, attr) in enumerate(EXPORT_COLUMNS, start=1):
                ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=_render_cell(getattr(d, attr)),
                )
            row_idx += 1

    # Reasonable column widths — cap at 40 so giant cells don't explode.
    for i, (h, attr) in enumerate(EXPORT_COLUMNS, start=1):
        widest = len(h)
        for g in result.groups:
            for d in g.devices:
                v = _render_cell(getattr(d, attr))
                if v and len(v) > widest:
                    widest = len(v)
        ws.column_dimensions[get_column_letter(i)].width = min(widest + 2, 40)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), f"{_filename_stub(result.tokens)}.xlsx"


def _render_cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(v)
    return str(v)
